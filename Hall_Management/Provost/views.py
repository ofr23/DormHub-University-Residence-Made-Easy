"""
Django views Configuration Documentation

This module defines various controller files
"""
from Student.models import *  # Import Student and Session models
from Varsity_Admin.models import *  # Import Hall and Room models
from django.contrib import messages
from django.shortcuts import render, redirect
from openpyxl import load_workbook

from .models import *


def provost(request):
    """
    View function to handle the provost view.

    This function retrieves the provost object based on the logged-in user and the associated hall.
    It then gets all rooms in the hall and creates a dictionary to store room availability information.
    The function processes allocation, adding student to room, and removing student from room form submissions.

    Args:
        request (HttpRequest): The HTTP request object.

    Returns:
        HttpResponse: The HTTP response object rendering the provost template.
    """
    # Get the provost object based on the logged-in user
    provost = Provost.objects.get(email=request.user.email)
    # Get the hall associated with the provost
    hall = Hall.objects.get(provost=provost)
    # Get all rooms in the hall
    rooms = Room.objects.filter(hall=hall)

    # Create a dictionary to store room availability information
    queryDict = {}
    for o in range(1, 8):
        queryDict[o] = []

    # Iterate over rooms to get room availability
    for room in rooms:
        ro = Room.objects.get(roomId=room.roomId, hall=hall)
        available = room.capacity - len(Student.objects.filter(hall=hall, room=ro))
        queryDict[int(room.roomId / 100)].append((room, available))

    # Process allocation form submission
    if 'allocate' in request.POST:
        room = Room.objects.get(roomId=int(request.POST.get('allocate')), hall=hall)
        residents = Student.objects.filter(room=room, hall=hall)
        available = room.capacity - len(Student.objects.filter(hall=hall, room=room))
        notAllocated = Student.objects.filter(room=None, hall=hall)
        messages.success(request, {'available': available, 'room': room.roomId, 'residents': residents,
                                   'notAllocated': notAllocated}, extra_tags='allocation')

    # Process adding student to room form submission
    if 'add' in request.POST:
        student = Student.objects.get(studentId=request.POST.get('select'))
        room = Room.objects.get(hall=hall, roomId=int(request.POST.get('room')))
        student.room = room
        student.save()
        return redirect('/provost')

    # Process removing student from room form submission
    if 'remove' in request.POST:
        student = Student.objects.get(studentId=request.POST.get('remove'))
        student.room = None
        student.save()
        return redirect('/provost')

    context = {
        'hall': hall,
        'queryDict': queryDict
    }
    return render(request, 'provost.html', context)


def addStudent(request):
    """
    View function to handle adding students.

    This function retrieves all sessions and processes form submission to add students.

    Args:
        request (HttpRequest): The HTTP request object.

    Returns:
        HttpResponse: The HTTP response object rendering the allStudent template.
    """
    # Get all sessions
    sessions = Session.objects.all()
    provost = Provost.objects.get(email=request.user.email)
    hall = Hall.objects.get(provost=provost)

    # Process form submission to add students
    if 'add' in request.POST:
        session = Session.objects.get(session=int(request.POST.get('add')))
        csv = session.csvFile
        wb = load_workbook(csv)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if int(row[4]) == hall.hallId:
                newStudent = Student(
                    studentId=int(row[0]),
                    session=session,
                    name=row[1],
                    hall=hall,
                    email=row[2],
                    studentType=row[3]
                )
                newStudent.save()
        return redirect('/provost')

    # Create a dictionary to store session availability information
    queryDict = {}
    for o in sessions:
        session = Session.objects.get(session=o.session)
        ifPresent = Student.objects.filter(session=session, hall=hall)
        if ifPresent:
            queryDict[o] = 1
        else:
            queryDict[o] = 0

    context = {
        'queryDict': queryDict
    }
    return render(request, 'allStudent.html', context)


def swapRequests(request):
    """
    View function to handle swap requests.

    This function retrieves swap requests associated with the provost's hall.
    It processes form submissions to accept or reject swap requests.

    Args:
        request (HttpRequest): The HTTP request object.

    Returns:
        HttpResponse: The HTTP response object rendering the swapRequests template.
    """
    provost = Provost.objects.get(email=request.user.email)
    hall = Hall.objects.get(provost=provost)
    requests = SwapRequest.objects.filter(hall=hall, status=0)
    if 'accept' in request.POST:
        req = SwapRequest.objects.get(hall=hall, student=Student.objects.get(studentId=request.POST.get('accept')))
        req.status = 1
        req.save()
        return redirect('/provost/swapRequests')
    if 'reject' in request.POST:
        req = SwapRequest.objects.get(hall=hall, student=Student.objects.get(studentId=request.POST.get('reject')))
        req.status = 2
        req.save()
        return redirect('/provost/swapRequests')

    context = {
        'requests': requests
    }
    return render(request, 'swapRequests.html', context)